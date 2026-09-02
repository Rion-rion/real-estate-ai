from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill


PROJECT_ROOT = Path(__file__).resolve().parent

INPUT_DIR = PROJECT_ROOT / "data" / "input"

OUTPUT_FILE = (
    INPUT_DIR
    / "contract_history.xlsx"
)


def create_template():

    INPUT_DIR.mkdir(
        parents=True,
        exist_ok=True,
    )

    columns = [
        "property_id",
        "city",
        "district_name",
        "area_m2",
        "floor_plan",
        "building_age",
        "structure",
        "renovation",
        "use",
        "city_planning",
        "coverage_ratio",
        "floor_area_ratio",
        "listing_date",
        "asking_price",
        "contract_date",
        "contract_price",
    ]

    sample = [
        [
            "A001",
            "足立区",
            "千住",
            65.2,
            "3LDK",
            12,
            "RC",
            "未改装",
            "住宅",
            "商業地域",
            80,
            400,
            "2026-04-01",
            45_000_000,
            "2026-05-21",
            42_000_000,
        ]
    ]

    df = pd.DataFrame(
        sample,
        columns=columns,
    )

    df.to_excel(
        OUTPUT_FILE,
        index=False,
        sheet_name="成約履歴",
    )

    style_excel()

    print(
        "成約日数学習用テンプレートを作成しました。"
    )

    print(
        f"保存先: {OUTPUT_FILE}"
    )


def style_excel():

    workbook = load_workbook(
        OUTPUT_FILE
    )

    worksheet = workbook[
        "成約履歴"
    ]

    header_fill = PatternFill(
        fill_type="solid",
        fgColor="1F4E78",
    )

    header_font = Font(
        color="FFFFFF",
        bold=True,
    )

    for cell in worksheet[1]:

        cell.fill = header_fill
        cell.font = header_font

        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
        )

    worksheet.freeze_panes = "A2"

    for column in worksheet.columns:

        max_length = 0

        letter = (
            column[0]
            .column_letter
        )

        for cell in column:

            if cell.value is None:
                continue

            max_length = max(
                max_length,
                len(str(cell.value)),
            )

        worksheet.column_dimensions[
            letter
        ].width = min(
            max(max_length + 2, 12),
            25,
        )

    workbook.save(
        OUTPUT_FILE
    )


def main():

    create_template()


if __name__ == "__main__":
    main()