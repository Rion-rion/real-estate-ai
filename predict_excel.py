from pathlib import Path
import json

import numpy as np
import pandas as pd

from catboost import CatBoostRegressor
from openpyxl import load_workbook
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
)


PROJECT_ROOT = Path(__file__).resolve().parent

MODEL_FILE = (
    PROJECT_ROOT
    / "models"
    / "price_model.cbm"
)

METRICS_FILE = (
    PROJECT_ROOT
    / "output"
    / "price_model_metrics.json"
)

INPUT_DIR = (
    PROJECT_ROOT
    / "data"
    / "input"
)

INPUT_FILE = (
    INPUT_DIR
    / "prediction_input.xlsx"
)

OUTPUT_DIR = (
    PROJECT_ROOT
    / "output"
)

OUTPUT_FILE = (
    OUTPUT_DIR
    / "price_predictions.xlsx"
)


def load_model():
    if not MODEL_FILE.exists():
        raise FileNotFoundError(
            f"モデルが見つかりません: {MODEL_FILE}"
        )

    if not METRICS_FILE.exists():
        raise FileNotFoundError(
            f"モデル情報が見つかりません: {METRICS_FILE}"
        )

    with open(
        METRICS_FILE,
        "r",
        encoding="utf-8",
    ) as file:
        model_info = json.load(file)

    feature_columns = model_info.get(
        "features",
        []
    )

    categorical_columns = model_info.get(
        "categorical_features",
        []
    )

    selected_feature_set = model_info.get(
        "selected_feature_set",
        "unknown",
    )

    if not feature_columns:
        raise RuntimeError(
            "モデルの特徴量情報がありません。"
        )

    model = CatBoostRegressor()

    model.load_model(
        str(MODEL_FILE)
    )

    print("モデル読み込み完了")
    print(
        f"特徴量セット: "
        f"{selected_feature_set}"
    )

    return (
        model,
        feature_columns,
        categorical_columns,
        model_info,
    )


def create_excel_template(
    feature_columns,
):
    INPUT_DIR.mkdir(
        parents=True,
        exist_ok=True,
    )

    template = {
        "property_id": [
            "A001",
            "A002",
            "A003",
        ],

        "city": [
            "足立区",
            "世田谷区",
            "港区",
        ],

        "district_name": [
            "千住",
            "三軒茶屋",
            "六本木",
        ],

        "area_m2": [
            65.2,
            55.0,
            70.0,
        ],

        "floor_plan": [
            "3LDK",
            "2LDK",
            "2LDK",
        ],

        "building_age": [
            12,
            8,
            5,
        ],

        "structure": [
            "RC",
            "RC",
            "RC",
        ],

        "renovation": [
            "未改装",
            "改装済",
            "未改装",
        ],

        "use": [
            "住宅",
            "住宅",
            "住宅",
        ],

        "city_planning": [
            "商業地域",
            "近隣商業地域",
            "商業地域",
        ],

        "coverage_ratio": [
            80,
            80,
            80,
        ],

        "floor_area_ratio": [
            400,
            300,
            600,
        ],

        "transaction_year": [
            2026,
            2026,
            2026,
        ],

        "transaction_quarter": [
            3,
            3,
            3,
        ],

        "asking_price": [
            45_000_000,
            60_000_000,
            120_000_000,
        ],
    }

    df = pd.DataFrame(
        template
    )

    for column in feature_columns:

        if column not in df.columns:
            df[column] = np.nan

    df.to_excel(
        INPUT_FILE,
        index=False,
        sheet_name="予測入力",
    )

    style_excel(
        INPUT_FILE,
        "予測入力",
    )

    print()
    print(
        "Excel入力テンプレートを作成しました。"
    )

    print(
        f"保存先: {INPUT_FILE}"
    )

    print()
    print(
        "Excelを編集して保存した後、"
        "もう一度実行してください。"
    )


def load_input(
    feature_columns,
):
    if not INPUT_FILE.exists():
        return None

    df = pd.read_excel(
        INPUT_FILE,
        sheet_name="予測入力",
    )

    if df.empty:
        raise RuntimeError(
            "Excelに予測対象がありません。"
        )

    missing_columns = [
        column
        for column in feature_columns
        if column not in df.columns
    ]

    if missing_columns:

        print()
        print(
            "不足カラム:"
        )

        for column in missing_columns:
            print(
                f"- {column}"
            )

        raise KeyError(
            "予測に必要な列が不足しています。"
        )

    print()
    print(
        f"予測対象: "
        f"{len(df):,}件"
    )

    return df


def prepare_input(
    df,
    feature_columns,
    categorical_columns,
):
    df = df.copy()

    if "area_m2" not in df.columns:
        raise KeyError(
            "area_m2 が必要です。"
        )

    for column in feature_columns:

        if column in categorical_columns:

            df[column] = (
                df[column]
                .fillna("不明")
                .astype(str)
            )

        else:

            df[column] = pd.to_numeric(
                df[column],
                errors="coerce",
            )

    df["area_m2"] = pd.to_numeric(
        df["area_m2"],
        errors="coerce",
    )

    if df["area_m2"].isna().any():

        raise ValueError(
            "area_m2 に入力ミスがあります。"
        )

    if (
        df["area_m2"] <= 0
    ).any():

        raise ValueError(
            "area_m2 は0より大きい"
            "数値にしてください。"
        )

    return df


def predict(
    model,
    df,
    feature_columns,
):
    X = df[
        feature_columns
    ]

    predicted_log_unit_price = (
        model.predict(X)
    )

    predicted_unit_price = (
        np.expm1(
            predicted_log_unit_price
        )
    )

    predicted_unit_price = np.maximum(
        predicted_unit_price,
        0,
    )

    predicted_contract_price = (
        predicted_unit_price
        * df["area_m2"].to_numpy()
    )

    return (
        predicted_unit_price,
        predicted_contract_price,
    )


def get_price_evaluation(
    gap_ratio,
):
    if pd.isna(gap_ratio):
        return ""

    if gap_ratio <= -0.10:
        return "割安"

    if gap_ratio <= 0.10:
        return "適正"

    if gap_ratio <= 0.20:
        return "やや割高"

    return "割高"


def create_result(
    df,
    predicted_unit_price,
    predicted_contract_price,
):
    result = df.copy()

    result[
        "AI予測㎡単価"
    ] = np.round(
        predicted_unit_price
    ).astype(int)

    result[
        "AI予測成約価格"
    ] = np.round(
        predicted_contract_price
    ).astype(int)

    if "asking_price" in result.columns:

        result[
            "asking_price"
        ] = pd.to_numeric(
            result[
                "asking_price"
            ],
            errors="coerce",
        )

        result[
            "売出価格との差額"
        ] = (
            result[
                "asking_price"
            ]
            - result[
                "AI予測成約価格"
            ]
        )

        result[
            "価格乖離率"
        ] = (
            result[
                "売出価格との差額"
            ]
            / result[
                "AI予測成約価格"
            ]
        )

        result[
            "価格評価"
        ] = (
            result[
                "価格乖離率"
            ]
            .apply(
                get_price_evaluation
            )
        )

    return result


def create_model_info(
    model_info,
):
    metrics = (
        model_info.get(
            "test_metrics"
        )
        or
        model_info.get(
            "metrics",
            {}
        ).get(
            "test",
            {}
        )
    )

    rows = [
        [
            "モデル",
            model_info.get(
                "model",
                "CatBoostRegressor",
            ),
        ],

        [
            "バージョン",
            model_info.get(
                "version",
                "",
            ),
        ],

        [
            "特徴量セット",
            model_info.get(
                "selected_feature_set",
                "",
            ),
        ],

        [
            "テストMAPE",
            metrics.get(
                "mape",
                "",
            ),
        ],

        [
            "テストR²",
            metrics.get(
                "r2",
                "",
            ),
        ],

        [
            "テストMAE",
            metrics.get(
                "mae",
                "",
            ),
        ],

        [
            "テストRMSE",
            metrics.get(
                "rmse",
                "",
            ),
        ],
    ]

    return pd.DataFrame(
        rows,
        columns=[
            "項目",
            "値",
        ],
    )


def save_excel(
    result,
    model_info,
):
    OUTPUT_DIR.mkdir(
        parents=True,
        exist_ok=True,
    )

    model_df = create_model_info(
        model_info
    )

    with pd.ExcelWriter(
        OUTPUT_FILE,
        engine="openpyxl",
    ) as writer:

        result.to_excel(
            writer,
            index=False,
            sheet_name="予測結果",
        )

        model_df.to_excel(
            writer,
            index=False,
            sheet_name="モデル情報",
        )

    style_excel(
        OUTPUT_FILE,
        "予測結果",
    )

    style_excel(
        OUTPUT_FILE,
        "モデル情報",
    )

    print()
    print(
        "Excel予測完了"
    )

    print(
        f"保存先: {OUTPUT_FILE}"
    )


def style_excel(
    file_path,
    sheet_name,
):
    workbook = load_workbook(
        file_path
    )

    worksheet = workbook[
        sheet_name
    ]

    header_fill = PatternFill(
        fill_type="solid",
        fgColor="1F4E78",
    )

    header_font = Font(
        color="FFFFFF",
        bold=True,
    )

    thin_border = Border(
        bottom=Side(
            style="thin",
            color="D9E2F3",
        )
    )

    for cell in worksheet[1]:

        cell.fill = header_fill
        cell.font = header_font

        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
        )

    worksheet.freeze_panes = "A2"

    for column_cells in worksheet.columns:

        max_length = 0

        column_letter = (
            column_cells[0]
            .column_letter
        )

        for cell in column_cells:

            if cell.value is None:
                continue

            length = len(
                str(cell.value)
            )

            max_length = max(
                max_length,
                length,
            )

            cell.border = (
                thin_border
            )

        worksheet.column_dimensions[
            column_letter
        ].width = min(
            max(
                max_length + 2,
                10,
            ),
            25,
        )

    header_map = {
        cell.value: cell.column
        for cell in worksheet[1]
    }

    money_columns = [
        "asking_price",
        "AI予測㎡単価",
        "AI予測成約価格",
        "売出価格との差額",
    ]

    for name in money_columns:

        column_number = (
            header_map.get(name)
        )

        if not column_number:
            continue

        for row in range(
            2,
            worksheet.max_row + 1,
        ):

            worksheet.cell(
                row=row,
                column=column_number,
            ).number_format = (
                '#,##0'
            )

    ratio_column = (
        header_map.get(
            "価格乖離率"
        )
    )

    if ratio_column:

        for row in range(
            2,
            worksheet.max_row + 1,
        ):

            worksheet.cell(
                row=row,
                column=ratio_column,
            ).number_format = (
                '0.00%'
            )

    workbook.save(
        file_path
    )


def show_result(
    result,
):
    print()
    print(
        "予測結果"
    )

    for index, row in (
        result.iterrows()
    ):

        print()
        print(
            f"物件 {index + 1}"
        )

        if "property_id" in result.columns:

            print(
                f"ID: "
                f"{row['property_id']}"
            )

        if "city" in result.columns:

            print(
                f"地域: "
                f"{row['city']}"
            )

        if "district_name" in result.columns:

            print(
                f"地区: "
                f"{row['district_name']}"
            )

        print(
            "AI予測成約価格: "
            f"{row['AI予測成約価格']:,.0f}"
            "円"
        )

        if (
            "asking_price"
            in result.columns
            and pd.notna(
                row["asking_price"]
            )
        ):

            print(
                "売出価格: "
                f"{row['asking_price']:,.0f}"
                "円"
            )

            print(
                "価格評価: "
                f"{row['価格評価']}"
            )


def main():
    print(
        "東京都中古マンション"
        "Excel成約価格予測"
    )

    (
        model,
        feature_columns,
        categorical_columns,
        model_info,
    ) = load_model()

    if not INPUT_FILE.exists():

        create_excel_template(
            feature_columns
        )

        return

    df = load_input(
        feature_columns
    )

    df = prepare_input(
        df,
        feature_columns,
        categorical_columns,
    )

    (
        predicted_unit_price,
        predicted_contract_price,
    ) = predict(
        model,
        df,
        feature_columns,
    )

    result = create_result(
        df,
        predicted_unit_price,
        predicted_contract_price,
    )

    save_excel(
        result,
        model_info,
    )

    show_result(
        result
    )


if __name__ == "__main__":
    main()